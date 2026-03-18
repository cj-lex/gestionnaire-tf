"""
Registre des Timbres Fiscaux — Application Flask
Étude de commissaire de justice

INTÉGRITÉ DU REGISTRE :
  Les opérations sensibles (suppression, annulation, modification) sont
  accessibles via la page /admin (mot de passe requis).
  Elles peuvent aussi être effectuées manuellement dans les fichiers JSON :
    - Suppression   : retirer le bloc du timbre dans timbres_{année}.json
                      + supprimer le PDF correspondant dans data/pdfs/
    - Annulation    : remettre statut="disponible", dossier=null,
                      date_utilisation=null dans le fichier de l'année concernée
  Ne jamais éditer un fichier JSON pendant qu'un import est en cours.
"""

import io, json, os, re, socket, sys, threading, uuid, webbrowser
from datetime import date
from pathlib import Path

from flask import (Flask, flash, make_response, redirect,
                   render_template_string, request, send_file,
                   send_from_directory, session, url_for)
from pypdf import PdfReader, PdfWriter
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Chemins réseau (fixés en dur)
# ---------------------------------------------------------------------------
RESEAU_DIR  = Path(r"\\SERVEUR\COMMUN\GESTION-TF")
DATA_DIR    = RESEAU_DIR / "data"
PDF_DIR     = DATA_DIR   / "pdfs"
JUSTIF_DIR  = DATA_DIR   / "justificatifs"
JUSTIF_FILE = DATA_DIR   / "justificatifs.json"
AUDIT_FILE  = DATA_DIR   / "audit.log"

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
MONTANT_TIMBRE  = 50.0
SEUIL_ALERTE    = 5

# Mot de passe admin : priorité variable d'environnement ADMIN_PASSWORD,
# puis fichier config.json dans le répertoire réseau, puis valeur par défaut.
def _charger_admin_password() -> str:
    env_pw = os.getenv("ADMIN_PASSWORD", "")
    if env_pw:
        return env_pw
    config_file = RESEAU_DIR / "config.json"
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8") as fh:
                cfg = json.load(fh)
                if cfg.get("admin_password"):
                    return cfg["admin_password"]
        except Exception:
            pass
    return "ACTIA1"  # valeur par défaut si aucune configuration trouvée

ADMIN_PASSWORD = _charger_admin_password()

# ---------------------------------------------------------------------------
# Flask
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = "timbre-fiscal-etude-2025-xK9m"
_lock = threading.RLock()  # RLock : réentrant, évite les deadlocks lors des appels imbriqués

# ---------------------------------------------------------------------------
# Couche données year-aware
# ---------------------------------------------------------------------------

def data_file(year: int) -> Path:
    return DATA_DIR / f"timbres_{year}.json"


def load_year(year: int) -> list:
    f = data_file(year)
    with _lock:
        if not f.exists():
            return []
        try:
            with open(f, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (json.JSONDecodeError, OSError) as exc:
            print(f"[DONNÉES] Erreur lecture {f}: {exc}", file=sys.stderr)
            return []


def save_year(year: int, data: list):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with _lock:
        with open(data_file(year), "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)


def annees_disponibles() -> list:
    if not DATA_DIR.exists():
        return []
    return sorted(
        [int(p.stem.split("_")[1]) for p in DATA_DIR.glob("timbres_*.json")],
        reverse=True,
    )


def load_all() -> list:
    result = []
    for year in sorted(annees_disponibles()):
        result.extend(load_year(year))
    return result


def load_justificatifs() -> list:
    with _lock:
        if not JUSTIF_FILE.exists():
            return []
        try:
            with open(JUSTIF_FILE, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (json.JSONDecodeError, OSError) as exc:
            print(f"[DONNÉES] Erreur lecture justificatifs: {exc}", file=sys.stderr)
            return []


def save_justificatifs(data: list):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with _lock:
        with open(JUSTIF_FILE, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)



def save_timbre(timbre: dict):
    year = int(timbre["date_achat"][:4])
    timbres = load_year(year)
    for i, t in enumerate(timbres):
        if t["id"] == timbre["id"]:
            timbres[i] = timbre
            break
    save_year(year, timbres)


# ---------------------------------------------------------------------------
# Nommage lisible des fichiers PDF
# ---------------------------------------------------------------------------

def _sanitize(s: str) -> str:
    """Rend une chaîne sûre pour un nom de fichier (remplace les caractères spéciaux)."""
    return re.sub(r"[^\w\-]", "-", s).strip("-")


def _new_pdf_path(year_dir: Path, date_achat: str, numero: str) -> tuple[str, Path]:
    """Retourne (chemin_relatif_depuis_PDF_DIR, chemin_absolu) pour un PDF timbre."""
    stem  = f"{date_achat}_{_sanitize(numero)}"
    fname = f"{stem}.pdf"
    full  = year_dir / fname
    n = 1
    while full.exists():
        fname = f"{stem}_{n}.pdf"
        full  = year_dir / fname
        n += 1
    return f"{year_dir.name}/{fname}", full


def _new_justif_path(year_dir: Path, date_achat: str, n: int) -> tuple[str, Path]:
    """Retourne (chemin_relatif_depuis_JUSTIF_DIR, chemin_absolu) pour un justificatif."""
    fname = f"{date_achat}_justificatif_{n}.pdf"
    return f"{year_dir.name}/{fname}", year_dir / fname


def migrer_nommage():
    """Migration one-shot : renomme les anciens PDFs UUID vers le format lisible.

    Les fichiers au format UUID (sans '/' dans le chemin) sont déplacés dans un
    sous-dossier par année et renommés avec la date et le numéro de timbre.
    Les fichiers déjà au nouveau format sont ignorés.
    """
    # ── Timbres ──────────────────────────────────────────────────────────────
    for year in annees_disponibles():
        timbres  = load_year(year)
        modifie  = False
        year_dir = PDF_DIR / str(year)
        for t in timbres:
            old = t.get("pdf", "")
            if old and "/" not in old and "\\" not in old:  # format UUID plat
                old_path = PDF_DIR / old
                if old_path.exists():
                    year_dir.mkdir(parents=True, exist_ok=True)
                    rel, new_path = _new_pdf_path(year_dir, t["date_achat"], t["numero"])
                    old_path.rename(new_path)
                    t["pdf"] = rel
                    modifie  = True
        if modifie:
            save_year(year, timbres)

    # ── Justificatifs ────────────────────────────────────────────────────────
    justifs  = load_justificatifs()
    modifie  = False
    compteurs: dict[str, int] = {}
    for j in justifs:
        old = j.get("pdf", "")
        if old and "/" not in old and "\\" not in old:
            old_path = JUSTIF_DIR / old
            if old_path.exists():
                date_a   = j["date_achat"]
                yr       = int(date_a[:4])
                year_dir = JUSTIF_DIR / str(yr)
                year_dir.mkdir(parents=True, exist_ok=True)
                compteurs[date_a] = compteurs.get(date_a, 0) + 1
                rel, new_path = _new_justif_path(year_dir, date_a, compteurs[date_a])
                old_path.rename(new_path)
                j["pdf"] = rel
                modifie  = True
    if modifie:
        save_justificatifs(justifs)


# ---------------------------------------------------------------------------
# Extraction du numéro de timbre
# ---------------------------------------------------------------------------

_PATTERNS = [
    re.compile(r"\b3[A-Z0-9]{2}[\s\-]?[0-9]{4}[\s\-]?[0-9]{4}[\s\-]?[0-9]{4}\b"),
    re.compile(r"\b[0-9]{4}[\s\-][0-9]{4}[\s\-][0-9]{4}[\s\-][0-9]{4}\b"),
    re.compile(r"\b[0-9]{13,20}\b"),
    re.compile(r"[Nn]°\s*([A-Z0-9\-]{8,})"),
    re.compile(r"[Rr][ée]f[ée]rence\s*:?\s*([A-Z0-9\-]{6,})"),
]


def extraire_numero(text: str) -> str | None:
    for pat in _PATTERNS:
        m = pat.search(text)
        if m:
            raw = m.group(1) if m.lastindex else m.group(0)
            return re.sub(r"\s+", "-", raw.strip())
    return None


_MOIS_FR = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "août": 8, "aout": 8,
    "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12, "decembre": 12,
}


def extraire_date(text: str) -> str | None:
    """Extrait une date au format ISO YYYY-MM-DD depuis le texte d'un PDF.

    Reconnaît les formats :
      - ISO        : 2026-03-17
      - Français   : 17/03/2026  17-03-2026  17.03.2026
      - Littéral   : 17 mars 2026
    Priorité aux dates situées près des mots-clés « paiement », « achat », « émis ».
    """
    def _chercher(src: str) -> str | None:
        # ISO
        m = re.search(r"\b(20\d{2})-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])\b", src)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        # Français numérique DD/MM/YYYY (séparateur / - .)
        m = re.search(r"\b(0[1-9]|[12]\d|3[01])[/\-\.](0[1-9]|1[0-2])[/\-\.](20\d{2})\b", src)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        # Français littéral : 17 mars 2026
        m = re.search(
            r"\b(0?[1-9]|[12]\d|3[01])\s+"
            r"(janvier|f[ée]vrier|mars|avril|mai|juin|juillet|ao[uû]t"
            r"|septembre|octobre|novembre|d[ée]cembre)\s+(20\d{2})\b",
            src, re.IGNORECASE,
        )
        if m:
            mois = _MOIS_FR.get(m.group(2).lower().replace("é", "e")
                                .replace("è", "e").replace("û", "u"), 0)
            if mois:
                return f"{m.group(3)}-{str(mois).zfill(2)}-{m.group(1).zfill(2)}"
        return None

    # Chercher d'abord près des mots-clés caractéristiques
    for mot in ("paiement", "achat", "émis", "emis", "payé", "paye", "date"):
        idx = text.lower().find(mot)
        if idx != -1:
            d = _chercher(text[max(0, idx - 10): idx + 80])
            if d:
                return d
    # Fallback : première date trouvée dans tout le texte
    return _chercher(text)


# ---------------------------------------------------------------------------
# Port disponible
# ---------------------------------------------------------------------------

def trouver_port(port_base: int = 5000, tentatives: int = 10) -> int:
    for port in range(port_base, port_base + tentatives):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(("127.0.0.1", port)) != 0:
                return port
    return port_base  # fallback


# ---------------------------------------------------------------------------
# Template de base
# ---------------------------------------------------------------------------

_BASE = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ title }} — Timbres Fiscaux</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=EB+Garamond:ital,wght@0,400;0,600;0,700;1,400&display=swap" rel="stylesheet">
<style>
:root{
  --marine:#1a2744;--marine2:#243156;--or:#c9a84c;--creme:#f5e6c8;
  --fond:#f7f5f0;--fond-row:#faf9f6;--bord:#e8e3d8;--bord2:#d0c8b8;
  --vbg:#d4edda;--vfg:#155724;--rbg:#f8d7da;--rfg:#721c24;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'EB Garamond',Georgia,serif;background:var(--fond);color:#2c2c2c;min-height:100vh}

/* ── NAV ── */
nav{background:linear-gradient(135deg,var(--marine),var(--marine2));border-bottom:3px solid var(--or);
    padding:0 2rem;display:flex;align-items:center;gap:2rem;height:58px}
.nav-brand{font-size:1.25rem;font-weight:700;color:var(--creme);text-decoration:none;white-space:nowrap;letter-spacing:.03em}
.nav-brand span{color:var(--or)}
.nav-links{display:flex;flex:1}
.nav-links a{color:#c8d4e8;text-decoration:none;padding:.4rem 1.1rem;font-size:1rem;
             border-bottom:3px solid transparent;margin-bottom:-3px;transition:color .2s,border-color .2s}
.nav-links a:hover{color:var(--creme)}
.nav-links a.active{color:var(--or);border-color:var(--or);background:rgba(201,168,76,.07)}
.nav-right{margin-left:auto}
.btn-excel{border:1.5px solid var(--or);color:var(--or);background:transparent;padding:.35rem .9rem;
           border-radius:6px;text-decoration:none;font-size:.92rem;font-family:inherit;cursor:pointer;transition:background .2s,color .2s}
.btn-excel:hover{background:var(--or);color:var(--marine)}

/* ── LAYOUT ── */
main{max-width:1100px;margin:0 auto;padding:2rem 1.5rem}
h1{font-size:1.7rem;color:var(--marine);margin-bottom:1.5rem;font-weight:600}
h2{font-size:1.2rem;color:var(--marine);margin-bottom:1rem;font-weight:600}

/* ── ALERTS / FLASH ── */
.alert{padding:.85rem 1.2rem;border-radius:8px;margin-bottom:1.2rem;font-size:1rem;border-left:5px solid}
.alert-danger {background:var(--rbg);color:var(--rfg);border-color:#e74c3c}
.alert-success{background:var(--vbg);color:var(--vfg);border-color:#27ae60}
.alert-error  {background:var(--rbg);color:var(--rfg);border-color:#e74c3c}

/* ── CARDS ── */
.card{background:#fff;border:1px solid var(--bord);border-radius:10px;
      padding:1.4rem 1.6rem;box-shadow:0 2px 8px rgba(26,39,68,.07);margin-bottom:1.4rem}

/* ── STATS ── */
.stats-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:1.2rem;margin-bottom:1.4rem}
.stat-card{background:#fff;border:1px solid var(--bord);border-radius:10px;
           padding:1.3rem 1.5rem;box-shadow:0 2px 8px rgba(26,39,68,.07);text-align:center}
.stat-card .lbl{font-size:.82rem;color:#666;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.5rem}
.stat-card .val{font-size:2.5rem;font-weight:700;line-height:1}
.stat-card.total .val{color:var(--marine)}
.stat-card.dispo .val{color:var(--vfg)}
.stat-card.used  .val{color:var(--rfg)}

/* ── WIDGETS ── */
.widgets-grid{display:grid;grid-template-columns:1fr 1fr;gap:1.2rem;margin-bottom:1.4rem}
.widget{background:#fff;border:1px solid var(--bord);border-left:4px solid var(--or);
        border-radius:10px;padding:1.2rem 1.5rem;box-shadow:0 2px 8px rgba(26,39,68,.07)}
.widget .wlbl{font-size:.78rem;text-transform:uppercase;letter-spacing:.07em;color:#999;margin-bottom:.45rem}
.widget .wmain{font-size:1.08rem;font-weight:600;color:var(--marine)}
.widget .wsub {font-size:.9rem;color:#666;margin-top:.2rem}

/* ── FORM ── */
label{display:block;font-size:.95rem;color:#444;margin-bottom:.3rem}
input[type=text],input[type=date]{width:100%;padding:.55rem .8rem;border:1.5px solid var(--bord2);
  border-radius:7px;font-family:inherit;font-size:1rem;background:var(--fond-row);transition:border-color .2s}
input[type=text]:focus,input[type=date]:focus{outline:none;border-color:var(--or)}
.form-row{margin-bottom:1rem}
.btn{background:var(--or);color:var(--marine);border:none;padding:.6rem 1.4rem;
     border-radius:7px;font-family:inherit;font-size:1rem;font-weight:700;cursor:pointer;transition:filter .2s}
.btn:hover{filter:brightness(1.08)}

/* ── DROPZONE ── */
.dropzone{border:2.5px dashed var(--or);border-radius:10px;background:#faf8f3;
          padding:2rem;text-align:center;cursor:pointer;transition:background .2s;margin-bottom:.5rem}
.dropzone:hover,.dropzone.dragover{background:#f5f0e0}
.dz-icon{font-size:2.5rem;margin-bottom:.5rem}
.dz-text{color:#666;font-size:.95rem}
.dz-price{color:var(--or);font-weight:600;margin-top:.4rem;font-size:.88rem}
#pdf-input{display:none}
#file-name{font-size:.88rem;color:var(--marine);margin-top:.4rem;font-style:italic}

/* ── TIMBRE FICHE ── */
.timbre-fiche{display:flex;align-items:stretch;background:#fff;border:1px solid var(--bord);
              border-radius:10px;box-shadow:0 2px 8px rgba(26,39,68,.07);margin-bottom:1.2rem;overflow:hidden}
.tf-cell{padding:1.1rem 1.5rem;border-right:1px solid var(--bord);flex:1}
.tf-cell:last-child{border-right:none}
.tf-lbl{font-size:.75rem;text-transform:uppercase;letter-spacing:.07em;color:#999;margin-bottom:.3rem}
.tf-val{font-size:1.05rem;font-weight:600;color:var(--marine)}
.tf-mono{font-family:'Courier New',monospace;font-size:1.1rem;font-weight:700;
         color:var(--marine);letter-spacing:.05em}

/* ── TABLE ── */
table{width:100%;border-collapse:collapse;font-size:.97rem}
thead tr{background:var(--marine);color:var(--creme)}
thead th{padding:.75rem 1rem;text-align:left;font-weight:600;
         border-bottom:2px solid var(--or);letter-spacing:.03em}
tbody tr{border-bottom:1px solid var(--bord)}
tbody tr:nth-child(odd) {background:#fff}
tbody tr:nth-child(even){background:var(--fond-row)}
tbody td{padding:.65rem 1rem;vertical-align:middle}

/* ── BADGES ── */
.badge{display:inline-block;padding:.2rem .7rem;border-radius:50px;font-size:.82rem;font-weight:600}
.badge-d{background:var(--vbg);color:var(--vfg)}
.badge-u{background:var(--rbg);color:var(--rfg)}
.mono{font-family:'Courier New',monospace;font-size:.93rem;letter-spacing:.04em}

/* ── LOT BLOCK ── */
.lot-block{margin-bottom:1.8rem;border-radius:8px;overflow:hidden;border:1px solid var(--bord)}
.lot-header{background:var(--marine);color:var(--creme);padding:.8rem 1.1rem;
            border-bottom:2px solid var(--or);font-size:.96rem;
            display:flex;align-items:center;gap:1rem;flex-wrap:wrap}
.lot-pill{background:rgba(201,168,76,.25);color:var(--or);border:1px solid rgba(201,168,76,.5);
          padding:.15rem .6rem;border-radius:50px;font-size:.82rem;font-weight:600;white-space:nowrap}

/* ── TOOLBAR ── */
.toolbar{display:flex;gap:1rem;align-items:center;margin-bottom:1.2rem;flex-wrap:wrap}
.toolbar input[type=text]{max-width:280px}
select{padding:.5rem .8rem;border:1.5px solid var(--bord2);border-radius:7px;
       font-family:inherit;font-size:1rem;background:var(--fond-row);cursor:pointer}
select:focus{outline:none;border-color:var(--or)}

/* ── MODAL PDF ── */
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;
               align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal-box{background:#fff;border-radius:10px;width:min(850px,95vw);height:min(680px,90vh);
           display:flex;flex-direction:column;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,.35)}
.modal-bar{background:var(--marine);color:var(--creme);padding:.7rem 1rem;
           display:flex;align-items:center;gap:.8rem;border-bottom:2px solid var(--or)}
.modal-bar .modal-title{flex:1;font-weight:600;font-size:.97rem}
.modal-bar a.btn-dl{font-size:.85rem;padding:.3rem .8rem;background:var(--or);
                    color:var(--marine);border:none;border-radius:5px;text-decoration:none;font-weight:700}
.modal-bar button.close-btn{background:transparent;border:none;color:var(--creme);
                             font-size:1.3rem;cursor:pointer;line-height:1;padding:0 .3rem}
.modal-iframe{flex:1;border:none;width:100%;background:#555}

/* ── SPINNER PAGE ── */
.spin-wrap{display:flex;flex-direction:column;align-items:center;justify-content:center;
           min-height:60vh;gap:1.5rem}
.spinner{width:60px;height:60px;border:5px solid #e8e3d8;border-top-color:var(--or);
         border-radius:50%;animation:spin .8s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.spin-text{font-size:1.2rem;color:var(--marine)}
.spin-sub{color:#888;font-size:.9rem}

/* ── MISC ── */
.subtitle{color:#555;font-size:1rem;margin-bottom:1.2rem}
.note{font-size:.9rem;color:#777;margin-top:.8rem;font-style:italic}
.empty{text-align:center;padding:2.5rem;color:#888;font-size:1.05rem}
@media(max-width:700px){
  .stats-grid,.widgets-grid{grid-template-columns:1fr}
  .timbre-fiche{flex-direction:column}
  .tf-cell{border-right:none;border-bottom:1px solid var(--bord);width:100%}
}
</style>
</head>
<body>
<nav>
  <a href="/" class="nav-brand">⚖ Timbres <span>Fiscaux</span></a>
  <div class="nav-links">
    <a href="/" class="{{ 'active' if active=='db' else '' }}">Tableau de bord</a>
    <a href="/disponibles" class="{{ 'active' if active=='dispo' else '' }}">Disponibles</a>
    <a href="/historique" class="{{ 'active' if active=='hist' else '' }}">Historique</a>
    <a href="/justificatifs" class="{{ 'active' if active=='justif' else '' }}">Justificatifs</a>
    <a href="/admin" class="{{ 'active' if active=='admin' else '' }}">⚙ Administration</a>
  </div>
  <div class="nav-right" style="position:relative">
    <button class="btn-excel" id="excel-toggle" type="button">⬇ Excel ▾</button>
    <div id="excel-menu" style="display:none;position:absolute;right:0;top:calc(100% + 6px);background:#fff;border:1px solid #ddd;border-radius:8px;box-shadow:0 4px 16px rgba(0,0,0,.15);padding:.7rem 1rem;min-width:220px;z-index:200">
      <div style="font-size:.82rem;color:#888;margin-bottom:.5rem;text-transform:uppercase;letter-spacing:.05em">Filtres export</div>
      <form action="/export-excel" method="get">
        <select name="annee" style="width:100%;margin-bottom:.5rem;font-size:.9rem">
          <option value="">Toutes les années</option>
          {% for a in annees %}<option value="{{ a }}">{{ a }}</option>{% endfor %}
        </select>
        <select name="statut" style="width:100%;margin-bottom:.7rem;font-size:.9rem">
          <option value="">Tous les statuts</option>
          <option value="disponible">Disponibles</option>
          <option value="utilisé">Utilisés</option>
        </select>
        <button type="submit" class="btn" style="width:100%;font-size:.9rem">⬇ Télécharger</button>
      </form>
    </div>
  </div>
  <script>
  (function(){
    var btn=document.getElementById('excel-toggle');
    var menu=document.getElementById('excel-menu');
    btn.addEventListener('click',function(e){e.stopPropagation();menu.style.display=menu.style.display==='none'?'block':'none';});
    document.addEventListener('click',function(){menu.style.display='none';});
    menu.addEventListener('click',function(e){e.stopPropagation();});
  })();
  </script>
</nav>
<main>
{% with messages = get_flashed_messages(with_categories=true) %}
  {% for cat, msg in messages %}
    <div class="alert alert-{{ cat }}">{{ msg }}</div>
  {% endfor %}
{% endwith %}
{{ content | safe }}
</main>

<!-- Modal PDF -->
<div class="modal-overlay" id="pdf-modal" onclick="fermerModal(event)">
  <div class="modal-box">
    <div class="modal-bar">
      <span class="modal-title" id="modal-title">Timbre fiscal</span>
      <a class="btn-dl" id="modal-dl" href="#" download>⬇ Télécharger</a>
      <button class="close-btn" onclick="document.getElementById('pdf-modal').classList.remove('open')">✕</button>
    </div>
    <iframe class="modal-iframe" id="modal-frame" src=""></iframe>
  </div>
</div>
<script>
function ouvrirPdf(url, titre) {
  document.getElementById('modal-frame').src = url;
  document.getElementById('modal-title').textContent = titre;
  document.getElementById('modal-dl').href = url;
  document.getElementById('modal-dl').download = titre + '.pdf';
  document.getElementById('pdf-modal').classList.add('open');
}
function fermerModal(e) {
  if (e.target === document.getElementById('pdf-modal')) {
    document.getElementById('pdf-modal').classList.remove('open');
    document.getElementById('modal-frame').src = '';
  }
}
document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') {
    document.getElementById('pdf-modal').classList.remove('open');
    document.getElementById('modal-frame').src = '';
  }
});
</script>
</body>
</html>"""


def render_page(title: str, active: str, content: str) -> str:
    return render_template_string(
        _BASE, title=title, active=active, content=content,
        annees=annees_disponibles(),
    )


# ---------------------------------------------------------------------------
# PAGE 1 — TABLEAU DE BORD
# ---------------------------------------------------------------------------

@app.route("/")
def dashboard():
    timbres   = load_all()
    dispos    = [t for t in timbres if t["statut"] == "disponible"]
    utilises  = [t for t in timbres if t["statut"] == "utilisé"]
    nb_dispo  = len(dispos)
    nb_used   = len(utilises)
    nb_total  = len(timbres)

    # Widget dernier import
    dernier_import = None
    if timbres:
        date_max = max(t["date_achat"] for t in timbres)
        lot = [t for t in timbres if t["date_achat"] == date_max]
        dernier_import = {"date": date_max, "nb": len(lot)}

    # Widget dernière attribution
    derniere_attrib = None
    if utilises:
        derniere_attrib = max(
            utilises, key=lambda t: t.get("date_utilisation") or ""
        )

    today = date.today().isoformat()

    alerte = ""
    if nb_dispo <= SEUIL_ALERTE:
        alerte = (
            f'<div class="alert alert-danger">⚠ Stock bas — {nb_dispo} timbre(s) disponible(s) · '
            f'Pensez à commander un nouveau lot (seuil d\'alerte : {SEUIL_ALERTE} timbres)</div>'
        )

    wi_import = '<div class="wmain" style="color:#aaa">Aucun import</div>'
    if dernier_import:
        wi_import = (
            f'<div class="wmain">{dernier_import["date"]} · {dernier_import["nb"]} timbre(s)</div>'
            f'<div class="wsub">{MONTANT_TIMBRE:.0f} € / timbre</div>'
        )

    wi_attrib = '<div class="wmain" style="color:#aaa">Aucune attribution</div>'
    if derniere_attrib:
        wi_attrib = (
            f'<div class="wmain mono">{derniere_attrib["numero"]}</div>'
            f'<div class="wsub">{derniere_attrib.get("dossier") or ""}</div>'
            f'<div class="wsub">{derniere_attrib.get("date_utilisation") or ""}</div>'
        )

    content = f"""{alerte}
<h1>Tableau de bord</h1>

<div class="stats-grid">
  <div class="stat-card total"><div class="lbl">Total</div><div class="val">{nb_total}</div></div>
  <div class="stat-card dispo"><div class="lbl">Disponibles</div><div class="val">{nb_dispo}</div></div>
  <div class="stat-card used"> <div class="lbl">Utilisés</div><div class="val">{nb_used}</div></div>
</div>

<div class="widgets-grid">
  <div class="widget"><div class="wlbl">Dernier import</div>{wi_import}</div>
  <div class="widget"><div class="wlbl">Dernière attribution</div>{wi_attrib}</div>
</div>

<div class="card">
  <h2>Importer un lot de timbres</h2>
  <form method="post" action="/import" enctype="multipart/form-data">
    <div class="form-row">
      <label for="date_achat">Date d'achat <span style="color:#999;font-size:.88rem;font-weight:400">(auto-détectée depuis le PDF si laissée vide)</span></label>
      <input type="date" id="date_achat" name="date_achat" value="" style="max-width:220px">
    </div>
    <div class="form-row">
      <label>Fichier PDF du lot</label>
      <div class="dropzone" id="dz" onclick="document.getElementById('pdf-input').click()">
        <div class="dz-icon">📄</div>
        <div class="dz-text">Glissez le PDF ici ou <strong>cliquez pour parcourir</strong></div>
        <div class="dz-price">PDF multi-pages · 1 page = 1 timbre · 50,00 € / timbre</div>
      </div>
      <input type="file" id="pdf-input" name="pdf" accept="application/pdf" required>
      <div id="file-name"></div>
    </div>
    <button type="submit" class="btn">Importer le lot</button>
  </form>
</div>

<script>
const dz=document.getElementById('dz'),inp=document.getElementById('pdf-input'),fn=document.getElementById('file-name');
inp.addEventListener('change',()=>{{fn.textContent=inp.files[0]?'📎 '+inp.files[0].name:''}});
dz.addEventListener('dragover',e=>{{e.preventDefault();dz.classList.add('dragover')}});
dz.addEventListener('dragleave',()=>dz.classList.remove('dragover'));
dz.addEventListener('drop',e=>{{e.preventDefault();dz.classList.remove('dragover');
  inp.files=e.dataTransfer.files;fn.textContent=inp.files[0]?'📎 '+inp.files[0].name:''}});
</script>"""

    return render_page("Tableau de bord", "db", content)


# ---------------------------------------------------------------------------
# IMPORT LOT
# ---------------------------------------------------------------------------

@app.route("/import", methods=["POST"])
def import_lot():
    pdf_file   = request.files.get("pdf")
    date_achat = request.form.get("date_achat", "").strip()

    if not pdf_file:
        flash("Fichier PDF requis.", "error")
        return redirect(url_for("dashboard"))

    try:
        reader = PdfReader(pdf_file.stream)

        # ── Extraction automatique de la date depuis le justificatif ──────────
        date_achat_pdf = None
        for page in reader.pages:
            text = page.extract_text() or ""
            if "JUSTIFICATIF DE PAIEMENT" in text.upper():
                date_achat_pdf = extraire_date(text)
                break  # une seule page justificatif suffit

        # Priorité : date extraite du PDF > date saisie manuellement
        date_achat_final = date_achat_pdf or date_achat
        if not date_achat_final:
            flash("Date d'achat introuvable dans le PDF. Veuillez la saisir manuellement.", "error")
            return redirect(url_for("dashboard"))

        if date_achat_pdf and date_achat_pdf != date_achat:
            flash(f"✓ Date d'achat auto-détectée depuis le PDF : {date_achat_pdf}.", "success")

        date_achat = date_achat_final
        year   = int(date_achat[:4])
        existing  = load_year(year)
        start_idx = len(existing)

        # Sous-dossiers par année
        year_pdf_dir    = PDF_DIR    / str(year)
        year_justif_dir = JUSTIF_DIR / str(year)
        year_pdf_dir.mkdir(parents=True, exist_ok=True)
        year_justif_dir.mkdir(parents=True, exist_ok=True)

        # Compteur de justificatifs pour cette date (pour le nommage)
        justif_n = len([j for j in load_justificatifs() if j["date_achat"] == date_achat])

        nouveaux = []

        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            # Extraire et sauvegarder la page "Justificatif de paiement à conserver"
            if "JUSTIFICATIF DE PAIEMENT" in text.upper():
                justif_n += 1
                justif_rel, justif_path = _new_justif_path(year_justif_dir, date_achat, justif_n)
                w = PdfWriter()
                w.add_page(page)
                with open(justif_path, "wb") as fj:
                    w.write(fj)
                justifs = load_justificatifs()
                justifs.append({
                    "id":         str(uuid.uuid4()),
                    "date_achat": date_achat,
                    "pdf":        justif_rel,
                })
                save_justificatifs(justifs)
                continue
            numero = extraire_numero(text) or f"TIMBRE-{date_achat}-{start_idx + i + 1:03d}"

            pdf_rel, pdf_path = _new_pdf_path(year_pdf_dir, date_achat, numero)
            writer = PdfWriter()
            writer.add_page(page)
            with open(pdf_path, "wb") as fout:
                writer.write(fout)

            nouveaux.append({
                "id":               str(uuid.uuid4()),
                "numero":           numero,
                "date_achat":       date_achat,
                "montant":          MONTANT_TIMBRE,
                "statut":           "disponible",
                "pdf":              pdf_rel,
                "dossier":          None,
                "date_utilisation": None,
            })

        existing.extend(nouveaux)
        save_year(year, existing)
        audit("IMPORT", f"{len(nouveaux)} timbre(s) importé(s) — date_achat={date_achat}")
        flash(f"✓ {len(nouveaux)} timbre(s) importé(s) avec succès ({year}).", "success")

    except Exception as exc:
        flash(f"Erreur lors de l'import : {exc}", "error")

    return redirect(url_for("dashboard"))


# ---------------------------------------------------------------------------
# PAGE 2 — DISPONIBLES
# ---------------------------------------------------------------------------

@app.route("/disponibles")
def disponibles():
    timbres = load_all()
    dispos  = sorted(
        [t for t in timbres if t["statut"] == "disponible"],
        key=lambda t: (t["date_achat"], t["id"]),
    )
    nb_dispo = len(dispos)
    prochain = dispos[0] if dispos else None

    if prochain:
        en_attente = nb_dispo - 1
        fiche = f"""
<div class="timbre-fiche">
  <div class="tf-cell">
    <div class="tf-lbl">N° Timbre</div>
    <div class="tf-mono">{prochain['numero']}</div>
  </div>
  <div class="tf-cell">
    <div class="tf-lbl">Montant</div>
    <div class="tf-val">{prochain['montant']:.2f} €</div>
  </div>
  <div class="tf-cell">
    <div class="tf-lbl">Date d'achat</div>
    <div class="tf-val">{prochain['date_achat']}</div>
  </div>
  <div class="tf-cell">
    <div class="tf-lbl">File</div>
    <div class="tf-val">1 / {nb_dispo}</div>
  </div>
</div>

<div class="card">
  <h2>Attribuer ce timbre</h2>
  <form method="post" action="/utiliser">
    <input type="hidden" name="timbre_id" value="{prochain['id']}">
    <div style="display:grid;grid-template-columns:1fr 180px;gap:1rem">
      <div class="form-row" style="margin-bottom:0">
        <label for="dossier">Référence dossier / affaire</label>
        <input type="text" id="dossier" name="dossier" autofocus required
               placeholder="Ex. : D250100, R10200, …">
      </div>
      <div class="form-row" style="margin-bottom:0">
        <label for="code_clerc">Code clerc</label>
        <input type="text" id="code_clerc" name="code_clerc" required
               placeholder="Ex. : JC, …">
      </div>
    </div>
    <button type="submit" class="btn" style="margin-top:1rem">Attribuer ce timbre</button>
  </form>
</div>
{"<p class='note'>📋 " + str(en_attente) + " timbre(s) en attente derrière celui-ci.</p>" if en_attente > 0 else ""}
"""
        subtitle = f"{nb_dispo} timbre(s) en stock · le suivant sera disponible après attribution"
    else:
        fiche    = "<div class='empty'>📭 Aucun timbre disponible. Importez un lot depuis le tableau de bord.</div>"
        subtitle = "Aucun timbre en stock"

    content = f"""<h1>Timbres disponibles</h1>
<p class="subtitle">{subtitle}</p>
{fiche}"""
    return render_page("Disponibles", "dispo", content)


# ---------------------------------------------------------------------------
# ATTRIBUTION
# ---------------------------------------------------------------------------

def _valider_dossier(dossier: str) -> str | None:
    """Retourne un message d'erreur si le dossier est invalide, sinon None."""
    if not dossier:
        return "La référence dossier est obligatoire."
    if len(dossier) > 100:
        return "La référence dossier ne peut pas dépasser 100 caractères."
    if not re.match(r'^[\w\s\-\./,()]+$', dossier):
        return "La référence dossier contient des caractères non autorisés."
    return None


def _valider_code_clerc(code: str) -> str | None:
    """Retourne un message d'erreur si le code clerc est invalide, sinon None."""
    if not code:
        return "Le code clerc est obligatoire."
    if len(code) > 20:
        return "Le code clerc ne peut pas dépasser 20 caractères."
    if not re.match(r'^[A-Za-z0-9\-]+$', code):
        return "Le code clerc ne doit contenir que des lettres, chiffres ou tirets."
    return None


@app.route("/utiliser", methods=["POST"])
def utiliser():
    timbre_id  = request.form.get("timbre_id",  "").strip()
    dossier    = request.form.get("dossier",    "").strip()
    code_clerc = request.form.get("code_clerc", "").strip()

    if not timbre_id or not dossier or not code_clerc:
        flash("Données manquantes.", "error")
        return redirect(url_for("disponibles"))

    err = _valider_dossier(dossier) or _valider_code_clerc(code_clerc)
    if err:
        flash(err, "error")
        return redirect(url_for("disponibles"))

    timbres = load_all()
    timbre  = next((t for t in timbres if t["id"] == timbre_id), None)

    if not timbre:
        flash("Timbre introuvable.", "error")
        return redirect(url_for("disponibles"))

    if timbre["statut"] != "disponible":
        flash("Ce timbre n'est plus disponible.", "error")
        return redirect(url_for("disponibles"))

    timbre["statut"]           = "utilisé"
    timbre["dossier"]          = dossier
    timbre["code_clerc"]       = code_clerc
    timbre["date_utilisation"] = date.today().isoformat()
    save_timbre(timbre)

    flash(f"✓ Timbre {timbre['numero']} attribué au dossier « {dossier} » (clerc : {code_clerc}).", "success")
    return redirect(url_for("telecharger", timbre_id=timbre_id))


@app.route("/attribution/telecharger/<timbre_id>")
def telecharger(timbre_id: str):
    timbres = load_all()
    timbre  = next((t for t in timbres if t["id"] == timbre_id), None)

    if not timbre:
        flash("Timbre introuvable.", "error")
        return redirect(url_for("disponibles"))

    pdf_url = url_for("serve_pdf", filename=timbre["pdf"])
    numero  = timbre["numero"]
    dossier = timbre.get("dossier") or ""

    content = f"""
<div class="spin-wrap">
  <div class="spinner"></div>
  <div class="spin-text">Téléchargement du timbre en cours…</div>
  <div class="spin-sub mono">{numero}</div>
  <div class="spin-sub">{dossier}</div>
</div>
<a id="dl" href="{pdf_url}" download="{numero}.pdf" style="display:none"></a>
<script>
window.onload=function(){{
  document.getElementById('dl').click();
  setTimeout(()=>window.location.href="{url_for('disponibles')}",1500);
}};
</script>"""
    return render_template_string(
        _BASE, title="Téléchargement", active="dispo", content=content
    )


# ---------------------------------------------------------------------------
# SERVE PDF (protégé : accessible uniquement après attribution)
# ---------------------------------------------------------------------------

@app.route("/pdfs/<path:filename>")
def serve_pdf(filename: str):
    timbres = load_all()
    timbre  = next((t for t in timbres if t.get("pdf") == filename), None)
    if not timbre or timbre["statut"] != "utilisé":
        return "Accès refusé.", 403
    full = (PDF_DIR / filename).resolve()
    if not str(full).startswith(str(PDF_DIR.resolve())):
        return "Accès refusé.", 403
    return send_file(str(full), mimetype="application/pdf")


# ---------------------------------------------------------------------------
# PAGE 3 — HISTORIQUE
# ---------------------------------------------------------------------------

LOTS_PAR_PAGE = 20


@app.route("/historique")
def historique():
    annee_param = request.args.get("annee", "toutes")
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    annees      = annees_disponibles()

    if annee_param == "toutes":
        timbres = load_all()
    else:
        try:
            timbres = load_year(int(annee_param))
        except ValueError:
            timbres = load_all()
            annee_param = "toutes"

    utilises = [t for t in timbres if t["statut"] == "utilisé"]

    # Regroupement par lot (date_achat, montant)
    lots: dict[tuple, list] = {}
    for t in utilises:
        lots.setdefault((t["date_achat"], t["montant"]), []).append(t)
    for k in lots:
        lots[k].sort(key=lambda t: t.get("date_utilisation") or "", reverse=True)
    lots_sorted = sorted(lots.items(), key=lambda kv: kv[0][0], reverse=True)

    def total_lot(date_a, montant):
        yr = int(date_a[:4])
        return sum(1 for t in load_year(yr)
                   if t["date_achat"] == date_a and t["montant"] == montant)

    nb_total   = len(utilises)
    mt_total   = nb_total * MONTANT_TIMBRE
    nb_lots    = len(lots)

    # Pagination
    nb_pages   = max(1, (nb_lots + LOTS_PAR_PAGE - 1) // LOTS_PAR_PAGE)
    page       = min(page, nb_pages)
    debut      = (page - 1) * LOTS_PAR_PAGE
    fin        = debut + LOTS_PAR_PAGE
    lots_page  = lots_sorted[debut:fin]

    # Sélecteur années
    opts  = f'<option value="toutes"{"  selected" if annee_param=="toutes" else ""}>Toutes les années</option>\n'
    for a in annees:
        sel   = ' selected' if str(a) == annee_param else ""
        opts += f'<option value="{a}"{sel}>{a}</option>\n'

    # Blocs de lots (page courante uniquement)
    blocs = ""
    for (date_a, montant), items in lots_page:
        nb_u   = len(items)
        tot    = total_lot(date_a, montant)
        mt_lot = nb_u * montant
        rows   = ""
        for t in items:
            if t.get("pdf"):
                btn_pdf = (
                    f'<div style="display:flex;gap:.4rem">'
                    f'<button class="btn" style="padding:.25rem .7rem;font-size:.82rem" '
                    f'onclick="ouvrirPdf(\'/pdfs/{t["pdf"]}\',\'{t["numero"]}\')">📄 Voir</button>'
                    f'<a class="btn" style="padding:.25rem .7rem;font-size:.82rem;text-decoration:none" '
                    f'href="/pdfs/{t["pdf"]}" download="timbre-{t["numero"]}.pdf">⬇ Télécharger</a>'
                    f'</div>'
                )
            else:
                btn_pdf = "—"
            rows += (
                f'<tr class="hr">'
                f'<td class="mono">{t["numero"]}</td>'
                f'<td>{t.get("dossier") or "—"}</td>'
                f'<td>{t.get("code_clerc") or "—"}</td>'
                f'<td>{t.get("date_utilisation") or "—"}</td>'
                f'<td>{btn_pdf}</td>'
                f'</tr>'
            )
        blocs += f"""
<div class="lot-block" data-lot>
  <div class="lot-header">
    📦 Lot du {date_a} · {montant:.2f} € / timbre
    <span class="lot-pill">{nb_u} utilisé(s) / {tot} dans ce lot</span>
    <span style="margin-left:auto;font-weight:600">{mt_lot:,.2f} €</span>
  </div>
  <table>
    <thead><tr><th>N° Timbre</th><th>Dossier / Affaire</th><th>Code clerc</th><th>Date utilisation</th><th>PDF</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
</div>"""

    if not blocs:
        sfx   = f" pour {annee_param}" if annee_param != "toutes" else ""
        blocs = f"<div class='empty'>Aucun timbre utilisé{sfx}.</div>"

    # Navigation pagination
    def _lien_page(p):
        return f"/historique?annee={annee_param}&page={p}"

    nav_pages = ""
    if nb_pages > 1:
        prev_btn = (
            f'<a href="{_lien_page(page-1)}" class="btn" style="padding:.35rem .8rem;font-size:.9rem;text-decoration:none">‹ Précédent</a>'
            if page > 1 else
            '<span class="btn" style="padding:.35rem .8rem;font-size:.9rem;opacity:.4;cursor:default">‹ Précédent</span>'
        )
        next_btn = (
            f'<a href="{_lien_page(page+1)}" class="btn" style="padding:.35rem .8rem;font-size:.9rem;text-decoration:none">Suivant ›</a>'
            if page < nb_pages else
            '<span class="btn" style="padding:.35rem .8rem;font-size:.9rem;opacity:.4;cursor:default">Suivant ›</span>'
        )
        nav_pages = f'<div style="display:flex;align-items:center;gap:1rem;margin-top:1.2rem">{prev_btn}<span style="color:#666;font-size:.92rem">Page {page} / {nb_pages}</span>{next_btn}</div>'

    subtitle = f"{nb_total} timbre(s) utilisé(s) sur {nb_lots} lot(s) — total : {mt_total:,.2f} €"

    content = f"""<h1>Historique des attributions</h1>
<p class="subtitle">{subtitle}</p>

<div class="toolbar">
  <input type="text" id="search" placeholder="Rechercher…" oninput="filtrer()">
  <select id="filtre-annee" onchange="window.location='/historique?annee='+this.value">
    {opts}
  </select>
</div>

{blocs}
{nav_pages}

<script>
function filtrer(){{
  const q=document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('[data-lot]').forEach(function(bloc){{
    let v=0;
    bloc.querySelectorAll('tr.hr').forEach(function(row){{
      const ok=row.textContent.toLowerCase().includes(q);
      row.style.display=ok?'':'none';
      if(ok)v++;
    }});
    bloc.style.display=v>0?'':'none';
  }});
}}
</script>"""

    return render_page("Historique", "hist", content)


# ---------------------------------------------------------------------------
# PAGE 4 — JUSTIFICATIFS DE PAIEMENT
# ---------------------------------------------------------------------------

@app.route("/justificatifs")
def justificatifs():
    justifs = sorted(load_justificatifs(), key=lambda j: j["date_achat"], reverse=True)

    # Regroupement par date_achat
    lots: dict[str, list] = {}
    for j in justifs:
        lots.setdefault(j["date_achat"], []).append(j)

    # Comptage des timbres par date d'achat
    tous_timbres = load_all()
    timbres_par_date: dict[str, list] = {}
    for t in tous_timbres:
        timbres_par_date.setdefault(t["date_achat"], []).append(t)

    blocs = ""
    for date_a, items in lots.items():
        timbres_lot = timbres_par_date.get(date_a, [])
        nb_timbres  = len(timbres_lot)
        montant_lot = sum(t["montant"] for t in timbres_lot)
        rows = ""
        for idx, j in enumerate(items, 1):
            pdf_url = url_for("serve_justificatif", filename=j["pdf"])
            rows += (
                f'<tr>'
                f'<td>Justificatif {idx}</td>'
                f'<td>{date_a}</td>'
                f'<td>{nb_timbres} timbre(s)</td>'
                f'<td style="font-weight:600">{montant_lot:,.2f} €</td>'
                f'<td style="display:flex;gap:.5rem">'
                f'<button class="btn" style="padding:.25rem .7rem;font-size:.82rem" '
                f'onclick="ouvrirPdf(\'{pdf_url}\',\'Justificatif {date_a}\')">📄 Voir</button>'
                f'<a class="btn" style="padding:.25rem .7rem;font-size:.82rem;text-decoration:none" '
                f'href="{pdf_url}" download="justificatif-{date_a}.pdf">⬇ Télécharger</a>'
                f'</td>'
                f'</tr>'
            )
        blocs += f"""
<div class="lot-block" style="margin-bottom:1.4rem">
  <div class="lot-header">
    🧾 Lot du {date_a}
    <span class="lot-pill">{len(items)} justificatif(s)</span>
    <span class="lot-pill">{nb_timbres} timbre(s) · {montant_lot:,.2f} €</span>
  </div>
  <table>
    <thead><tr><th>Document</th><th>Date d'achat</th><th>Nb timbres</th><th>Montant total</th><th>Actions</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
</div>"""

    if not blocs:
        blocs = "<div class='empty'>Aucun justificatif enregistré. Les prochains imports les ajouteront automatiquement.</div>"

    nb = len(justifs)
    content = f"""<h1>Justificatifs de paiement</h1>
<p class="subtitle">{nb} justificatif(s) conservé(s)</p>
{blocs}"""
    return render_page("Justificatifs", "justif", content)


@app.route("/justificatifs/pdf/<path:filename>")
def serve_justificatif(filename: str):
    justifs = load_justificatifs()
    if not any(j["pdf"] == filename for j in justifs):
        return "Accès refusé.", 403
    full = (JUSTIF_DIR / filename).resolve()
    if not str(full).startswith(str(JUSTIF_DIR.resolve())):
        return "Accès refusé.", 403
    return send_file(str(full), mimetype="application/pdf")


# ---------------------------------------------------------------------------
# EXPORT EXCEL
# ---------------------------------------------------------------------------

@app.route("/export-excel")
def export_excel():
    # Filtres optionnels : ?annee=2025&statut=utilisé
    filtre_annee  = request.args.get("annee", "").strip()
    filtre_statut = request.args.get("statut", "").strip().lower()

    annees = annees_disponibles()
    if not annees:
        flash("Aucune donnée à exporter.", "error")
        return redirect(url_for("dashboard"))

    # Restreindre aux années filtrées
    if filtre_annee:
        try:
            annee_int = int(filtre_annee)
            annees = [a for a in annees if a == annee_int]
        except ValueError:
            pass

    if not annees:
        flash(f"Aucune donnée pour l'année {filtre_annee}.", "error")
        return redirect(url_for("dashboard"))

    marine_hex = "1a2744"
    creme_hex  = "f5e6c8"
    or_hex     = "c9a84c"
    vbg, vfg   = "d4edda", "155724"
    rbg, rfg   = "f8d7da", "721c24"
    side_or    = Side(style="thin", color=or_hex)
    bord_or    = Border(left=side_or, right=side_or, top=side_or, bottom=side_or)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # supprimer l'onglet vide par défaut

    for year in sorted(annees, reverse=True):
        timbres_brut = load_year(year)
        # Appliquer le filtre statut si demandé
        if filtre_statut in ("disponible", "utilisé"):
            timbres_brut = [t for t in timbres_brut if t["statut"] == filtre_statut]
        timbres = sorted(timbres_brut, key=lambda t: (t["statut"], t["date_achat"]))
        if not timbres:
            continue
        ws = wb.create_sheet(title=str(year))

        headers = ["N° Timbre", "Date achat", "Montant (€)", "Statut", "Dossier", "Code clerc", "Date utilisation"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font      = Font(bold=True, color=creme_hex, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=marine_hex)
            c.border    = bord_or
            c.alignment = Alignment(horizontal="center", vertical="center")

        for t in timbres:
            label = "Disponible" if t["statut"] == "disponible" else "Utilisé"
            ws.append([
                t["numero"], t["date_achat"], t["montant"],
                label,
                t.get("dossier") or "",
                t.get("code_clerc") or "",
                t.get("date_utilisation") or "",
            ])
            r  = ws.max_row
            sc = ws.cell(row=r, column=4)
            if t["statut"] == "disponible":
                sc.fill = PatternFill("solid", fgColor=vbg)
                sc.font = Font(color=vfg, bold=True, name="Calibri")
            else:
                sc.fill = PatternFill("solid", fgColor=rbg)
                sc.font = Font(color=rfg, bold=True, name="Calibri")
            sc.alignment = Alignment(horizontal="center")
            # lignes alternées
            if r % 2 == 0:
                for col in [1, 2, 3, 5, 6]:
                    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="faf9f6")

        for i, w in enumerate([22, 15, 13, 14, 40, 14, 18], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22

    if not wb.sheetnames:
        flash("Aucune donnée à exporter avec les filtres sélectionnés.", "error")
        return redirect(url_for("dashboard"))

    buf      = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffixe  = f"-{filtre_annee}" if filtre_annee else ""
    suffixe += f"-{filtre_statut}" if filtre_statut else ""
    fname    = f"timbres-fiscaux{suffixe}-{date.today()}.xlsx"
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response


# ---------------------------------------------------------------------------
# PAGE 5 — ADMINISTRATION (mot de passe requis à chaque visite)
# ---------------------------------------------------------------------------

@app.route("/admin", methods=["GET", "POST"])
def admin():
    # Le mot de passe est exigé à chaque chargement de la page.
    # Un GET efface toujours la session et affiche le formulaire.
    # Un POST avec le bon mot de passe affiche le contenu admin.
    erreur = ""
    if request.method == "GET":
        # Effacer la session uniquement si l'utilisateur vient d'une page
        # extérieure à l'administration (navigation "aller-retour").
        # Si le Referer provient d'une URL admin (après une action interne),
        # la session est conservée pour éviter de redemander le mot de passe
        # après chaque modification.
        referer = request.headers.get("Referer", "")
        if "/admin" not in referer:
            session.pop("admin", None)
    elif request.method == "POST" and "password" in request.form:
        if request.form["password"] == ADMIN_PASSWORD:
            session["admin"] = True  # autorise les sous-actions POST
        else:
            erreur = "Mot de passe incorrect."

    if not session.get("admin") or erreur:
        session.pop("admin", None)
        content = f"""
<div style="max-width:380px;margin:4rem auto">
  <div class="card">
    <h2 style="margin-bottom:1.2rem">⚙ Administration</h2>
    {"<div class='alert alert-error'>" + erreur + "</div>" if erreur else ""}
    <form method="post">
      <div class="form-row">
        <label for="pw">Mot de passe</label>
        <input type="password" id="pw" name="password" autofocus required
               placeholder="••••••••" style="max-width:100%">
      </div>
      <button type="submit" class="btn" style="width:100%">Accéder</button>
    </form>
  </div>
</div>"""
        return render_page("Administration", "admin", content)

    annee_param = request.args.get("annee", "toutes")
    statut_param = request.args.get("statut", "tous")
    annees = annees_disponibles()

    if annee_param == "toutes":
        timbres = load_all()
    else:
        try:
            timbres = load_year(int(annee_param))
        except ValueError:
            timbres = load_all()
            annee_param = "toutes"

    if statut_param == "disponible":
        timbres = [t for t in timbres if t["statut"] == "disponible"]
    elif statut_param == "utilisé":
        timbres = [t for t in timbres if t["statut"] == "utilisé"]

    timbres_sorted = sorted(timbres, key=lambda t: (t["date_achat"], t.get("date_utilisation") or ""), reverse=True)

    # Sélecteur années
    opts_a = f'<option value="toutes"{"  selected" if annee_param=="toutes" else ""}>Toutes les années</option>\n'
    for a in annees:
        sel = " selected" if str(a) == annee_param else ""
        opts_a += f'<option value="{a}"{sel}>{a}</option>\n'

    opts_s = ""
    for val, lbl in [("tous","Tous les statuts"),("disponible","Disponibles"),("utilisé","Utilisés")]:
        sel = " selected" if statut_param == val else ""
        opts_s += f'<option value="{val}"{sel}>{lbl}</option>\n'

    rows = ""
    for t in timbres_sorted:
        badge = (
            "<span class='badge badge-d'>Disponible</span>"
            if t["statut"] == "disponible"
            else "<span class='badge badge-u'>Utilisé</span>"
        )
        dossier_val  = t.get("dossier") or ""
        clerc_val    = t.get("code_clerc") or ""
        date_achat_v = t.get("date_achat") or ""
        date_util    = t.get("date_utilisation") or date.today().isoformat()
        # Formulaire inline : dossier + code clerc + date achat + date utilisation
        form_attribution = f"""
<form method="post" action="/admin/modifier-dossier" style="display:flex;gap:.35rem;align-items:center;min-width:0">
  <input type="hidden" name="timbre_id" value="{t['id']}">
  <input type="text" name="dossier" value="{dossier_val}"
         style="flex:1 1 140px;min-width:0;padding:.3rem .5rem;font-size:.85rem"
         placeholder="Ex. : D250100, R10200" required>
  <input type="text" name="code_clerc" value="{clerc_val}"
         style="flex:0 0 65px;width:65px;padding:.3rem .5rem;font-size:.85rem"
         placeholder="Ex. : JC" required>
  <input type="date" name="date_achat" value="{date_achat_v}"
         title="Date d'achat"
         style="flex:0 0 122px;width:122px;padding:.3rem .4rem;font-size:.82rem">
  <input type="date" name="date_utilisation" value="{date_util}"
         title="Date d'utilisation"
         style="flex:0 0 122px;width:122px;padding:.3rem .4rem;font-size:.82rem">
  <button type="submit" class="btn" style="flex:0 0 auto;padding:.3rem .6rem;font-size:.82rem">✔</button>
</form>"""
        # Bouton remettre disponible (si utilisé)
        btn_reset = ""
        if t["statut"] == "utilisé":
            btn_reset = f"""
<form method="post" action="/admin/remettre-disponible" style="display:inline">
  <input type="hidden" name="timbre_id" value="{t['id']}">
  <button type="submit" class="btn"
          style="padding:.3rem .7rem;font-size:.82rem;background:#27ae60"
          onclick="return confirm('Remettre ce timbre en stock disponible ?')">↩ Disponible</button>
</form>"""
        # Bouton supprimer
        btn_suppr = f"""
<form method="post" action="/admin/supprimer" style="display:inline">
  <input type="hidden" name="timbre_id" value="{t['id']}">
  <button type="submit" class="btn"
          style="padding:.3rem .7rem;font-size:.82rem;background:#c0392b;color:#fff"
          onclick="return confirm('Supprimer définitivement ce timbre ? Cette action est irréversible.')">🗑 Supprimer</button>
</form>"""

        rows += f"""<tr class="adm-row">
  <td class="mono">{t['numero']}</td>
  <td>{t['date_achat']}</td>
  <td>{badge}</td>
  <td colspan="4">{form_attribution}</td>
  <td style="white-space:nowrap">{btn_reset} {btn_suppr}</td>
</tr>"""

    if not rows:
        rows = "<tr><td colspan='8' class='empty'>Aucun timbre trouvé.</td></tr>"

    content = f"""<h1>Administration</h1>
<p class="subtitle">{len(timbres_sorted)} timbre(s) affiché(s) &nbsp;·&nbsp;
<a href="/admin/lock" style="color:#c0392b;font-size:.9rem" title="Verrouiller l'administration">🔒 Verrouiller</a></p>

<div class="toolbar">
  <input type="text" id="search" placeholder="Rechercher…" oninput="filtrer()">
  <select id="f-annee" onchange="recharger()">
    {opts_a}
  </select>
  <select id="f-statut" onchange="recharger()">
    {opts_s}
  </select>
</div>

<div style="overflow-x:auto">
<table>
  <thead>
    <tr>
      <th>N° Timbre</th>
      <th>Date achat</th>
      <th>Statut</th>
      <th>Dossier / Affaire</th>
      <th>Code clerc</th>
      <th>Date achat</th>
      <th>Date utilisation</th>
      <th>Actions</th>
    </tr>
  </thead>
  <tbody id="tbody">{rows}</tbody>
</table>
</div>

<script>
function recharger(){{
  const a=document.getElementById('f-annee').value;
  const s=document.getElementById('f-statut').value;
  window.location='/admin?annee='+a+'&statut='+s;
}}
function filtrer(){{
  const q=document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('tr.adm-row').forEach(function(r){{
    r.style.display=r.textContent.toLowerCase().includes(q)?'':'none';
  }});
}}
</script>"""

    return render_page("Administration", "admin", content)


@app.route("/admin/modifier-dossier", methods=["POST"])
def admin_modifier_dossier():
    if not session.get("admin"):
        return "Accès refusé.", 403
    timbre_id  = request.form.get("timbre_id",  "").strip()
    dossier    = request.form.get("dossier",    "").strip()
    code_clerc = request.form.get("code_clerc", "").strip()

    err = _valider_dossier(dossier) or _valider_code_clerc(code_clerc)
    if err:
        flash(err, "error")
        return redirect(url_for("admin",
                                annee=request.args.get("annee", "toutes"),
                                statut=request.args.get("statut", "tous")))

    date_util      = request.form.get("date_utilisation", "").strip() or date.today().isoformat()
    date_achat_new = request.form.get("date_achat", "").strip()

    timbres = load_all()
    timbre  = next((t for t in timbres if t["id"] == timbre_id), None)
    if not timbre:
        flash("Timbre introuvable.", "error")
    else:
        year_old       = int(timbre["date_achat"][:4])
        date_achat_new = date_achat_new or timbre["date_achat"]
        year_new       = int(date_achat_new[:4])

        timbre["dossier"]          = dossier
        timbre["code_clerc"]       = code_clerc
        timbre["statut"]           = "utilisé"
        timbre["date_utilisation"] = date_util
        timbre["date_achat"]       = date_achat_new

        if year_old != year_new:
            # Déplacer le fichier PDF vers le sous-dossier de la nouvelle année
            if timbre.get("pdf"):
                old_pdf_full = PDF_DIR / timbre["pdf"]
                new_year_dir = PDF_DIR / str(year_new)
                new_year_dir.mkdir(parents=True, exist_ok=True)
                rel, new_pdf_full = _new_pdf_path(new_year_dir, date_achat_new, timbre["numero"])
                try:
                    old_pdf_full.rename(new_pdf_full)
                    timbre["pdf"] = rel
                except OSError as exc:
                    print(f"[PDF] Impossible de déplacer {old_pdf_full} → {new_pdf_full} : {exc}", file=sys.stderr)
                    flash(
                        f"Attention : le fichier PDF du timbre {timbre['numero']} n'a pas pu être "
                        f"déplacé vers l'année {year_new}. Veuillez le déplacer manuellement.",
                        "danger",
                    )
            # Retirer du fichier JSON de l'ancienne année
            old_timbres = load_year(year_old)
            save_year(year_old, [t for t in old_timbres if t["id"] != timbre_id])
            # Ajouter dans le fichier JSON de la nouvelle année
            new_timbres = load_year(year_new)
            new_timbres.append(timbre)
            save_year(year_new, new_timbres)
        else:
            save_timbre(timbre)

        flash(f"✓ Timbre {timbre['numero']} mis à jour (dossier : {dossier}, clerc : {code_clerc}).", "success")

    return redirect(url_for("admin",
                            annee=request.args.get("annee", "toutes"),
                            statut=request.args.get("statut", "tous")))


@app.route("/admin/lock")
def admin_lock():
    session.pop("admin", None)
    return redirect(url_for("admin"))


@app.route("/admin/remettre-disponible", methods=["POST"])
def admin_remettre_disponible():
    if not session.get("admin"):
        return "Accès refusé.", 403
    timbre_id = request.form.get("timbre_id", "").strip()

    timbres = load_all()
    timbre  = next((t for t in timbres if t["id"] == timbre_id), None)
    if not timbre:
        flash("Timbre introuvable.", "error")
    else:
        timbre["statut"]           = "disponible"
        timbre["dossier"]          = None
        timbre["date_utilisation"] = None
        save_timbre(timbre)
        audit("ADMIN_RESET", f"timbre={timbre['numero']} remis disponible")
        flash(f"✓ Timbre {timbre['numero']} remis en stock disponible.", "success")

    return redirect(url_for("admin"))


@app.route("/admin/supprimer", methods=["POST"])
def admin_supprimer():
    if not session.get("admin"):
        return "Accès refusé.", 403
    timbre_id = request.form.get("timbre_id", "").strip()

    timbres = load_all()
    timbre  = next((t for t in timbres if t["id"] == timbre_id), None)
    if not timbre:
        flash("Timbre introuvable.", "error")
        return redirect(url_for("admin"))

    # Supprimer le fichier PDF associé
    if timbre.get("pdf"):
        pdf_path = PDF_DIR / timbre["pdf"]
        try:
            pdf_path.unlink(missing_ok=True)
        except OSError:
            pass

    year = int(timbre["date_achat"][:4])
    timbres_annee = load_year(year)
    timbres_annee = [t for t in timbres_annee if t["id"] != timbre_id]
    save_year(year, timbres_annee)
    audit("ADMIN_SUPPRESSION", f"timbre={timbre['numero']} dossier={timbre.get('dossier') or '—'} date_achat={timbre['date_achat']}")
    flash(f"✓ Timbre {timbre['numero']} supprimé définitivement.", "success")

    return redirect(url_for("admin"))


# ---------------------------------------------------------------------------
# INTÉGRITÉ : aucune route de suppression ou d'annulation dans l'interface.
# Toute opération de ce type doit être effectuée manuellement dans les fichiers
# JSON (voir commentaire en tête de fichier et README.md).
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Vérification de l'accès au répertoire réseau
    try:
        PDF_DIR.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        print()
        print("  ERREUR — Impossible d'accéder au répertoire réseau :")
        print(f"  {PDF_DIR}")
        print()
        print(f"  Détail : {exc}")
        print()
        print("  Vérifiez que \\\\SERVEUR\\COMMUN\\GESTION-TF est bien accessible")
        print("  (lecteur réseau mappé ou UNC accessible).")
        print()
        input("  Appuyez sur Entrée pour quitter…")
        sys.exit(1)

    # Migration one-shot : renomme les anciens PDFs UUID en noms lisibles
    try:
        migrer_nommage()
    except Exception as exc:
        print(f"  Avertissement migration nommage : {exc}", file=sys.stderr)

    port   = trouver_port(5000)
    url    = f"http://localhost:{port}"
    annees = annees_disponibles()
    annees_str = ", ".join(str(a) for a in sorted(annees)) if annees else "aucune"

    print("=" * 60)
    print("  ⚖  Registre des Timbres Fiscaux")
    print("=" * 60)
    print(f"  Adresse      : {url}")
    print(f"  Données      : \\\\SERVEUR\\COMMUN\\GESTION-TF\\data")
    print(f"  PDFs         : \\\\SERVEUR\\COMMUN\\GESTION-TF\\data\\pdfs")
    print(f"  Années       : {annees_str}")
    print("=" * 60)
    print("  Fermer cette fenêtre pour quitter l'application.")
    print("=" * 60)

    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    app.run(host="127.0.0.1", port=port, debug=False)
